#!/usr/bin/env python3
"""
gather.py — Stage 1 of the Rights Left pipeline (GATHER).

Runs on a schedule (Sundays). Pulls the past week's political news from a set of
RSS feeds, keeps items that look like Trump-administration actions, and writes a
review workbook to  candidates/<monday>.xlsx  — a formatted Excel sheet with a
dropdown on the Include? column, so approving is done in Excel rather than by
hand-editing a raw CSV in the browser.

You then EDIT that workbook (the "approve" step) in Excel: pick "y" from the
Include? dropdown on every row you want kept (rows marked y highlight green as
you go). Category / event / impact are optional — leave them for enrich.py, or
fill them in yourself if you'd rather. Re-upload the edited file to candidates/
(replacing the original — same filename), then Stage 2 (ingest.py) appends the
approved rows into the tracker workbook automatically.

Optional: set an ANTHROPIC_API_KEY secret and pass --draft to have the model
pre-fill category / event / impact for you (you still review before ingest).
Without --draft, no API key is needed and nothing is sent anywhere.
"""

import os, sys, re, argparse, datetime as dt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter

try:
    import feedparser
except ImportError:
    sys.exit("feedparser not installed — run: pip install -r requirements.txt")

# ---- Feeds (edit freely; a feed that errors is skipped, not fatal) -----------
FEEDS = {
    "NPR":            "https://feeds.npr.org/1014/rss.xml",
    "CNN Politics":   "http://rss.cnn.com/rss/cnn_allpolitics.rss",
    "Politico":       "https://rss.politico.com/politics-news.xml",
    "NBC News":       "https://feeds.nbcnews.com/nbcnews/public/politics",
    "The Guardian":   "https://www.theguardian.com/us-news/us-politics/rss",
    "CBS News":       "https://www.cbsnews.com/latest/rss/politics",
    "ABC News":       "https://abcnews.go.com/abcnews/politicsheadlines",
    "PBS NewsHour":   "https://www.pbs.org/newshour/feeds/rss/politics",
}

# ---- Relevance filter: keep items whose title/summary mentions any of these --
KEYWORDS = [
    # --- Core administration & politics ---
    "trump", "white house", "executive order", "administration",
    "pardon", "vance", "rfk", "kennedy", "bondi", "patel", "epstein",
    "hegseth", "noem", "shutdown", "voting", "election", "federal",
    "insurrection", "national guard",

    # --- Immigration, ICE & DHS ---
    "ice ", "i.c.e.", "immigration", "deport", "deportation",
    "customs enforcement", "immigration enforcement",
    "ice agent", "ice raid", "ice arrest", "ice detention",
    "homeland security", "dhs", "border", "border patrol",
    "asylum", "refugee", "migrant", "visa", "birthright",
    "tps", "temporary protected status", "sanctuary",

    # --- Courts & law ---
    "supreme court", "scotus", "doj", "justice dept",
    "justice department",

    # --- Economy ---
    "tariff", "medicaid", "snap",

    # --- Science & public health ---
    "cdc", "nih", "fda", "vaccine", "public health",
    "climate", "environmental protection", "epa",
    "science", "scientific", "research funding",
    "anti-science", "rfk health",

    # --- Education ---
    "education department", "dept of education", "title ix",
    "student loan", "school", "university", "college",
    "academic freedom", "book ban", "curriculum",

    # --- Press & communications ---
    "fcc", "press freedom", "journalist",
]

# ---- The 17 categories the workbook uses (for the AI drafter / your reference)
CATEGORIES = [
    "Civil Liberties", "Civil Rights & Minorities", "Courts & SCOTUS",
    "Democracy & Rule of Law", "Economy & Tariffs", "Education",
    "Environment & Science", "Executive Power", "Federal Workforce",
    "Foreign Policy & Aid", "Healthcare", "Immigration",
    "Immigration / Free Speech", "LGBTQ+ Rights", "Press Freedom",
    "Public Health", "Women's Rights / LGBTQ+",
]

MAX_CANDIDATES = 60

# ---- Review workbook layout ---------------------------------------------------
# (row-dict key, column header, column width). Include + Headline lead since
# those are the two things you actually look at to decide keep-or-skip; the
# AI-optional fields are grouped together after; week_of/dates trail since
# they're administrative and rarely need a glance.
REVIEW_COLS = [
    ("include",  "Include?",                  10),
    ("srcdesc",  "Headline / Description",    55),
    ("outlet",   "Outlet",                    16),
    ("srcdate",  "Date",                      12),
    ("url",      "Link",                      40),
    ("category", "Category (optional)",       26),
    ("event",    "Event (optional)",          40),
    ("impact",   "Impact (optional)",         40),
    ("week_of",  "Week Of",                   14),
    ("dates",    "Date(s)",                   12),
]

NAVY, LIGHT_GREEN = "1F3864", "C6E8C6"
HDR_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
HDR_FILL = PatternFill("solid", start_color=NAVY)
BODY = Font(name="Arial", size=10)
LINK_FONT = Font(name="Arial", size=10, color="0563C1", underline="single")
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Border(bottom=Side(style="thin", color="BFBFBF"))


def write_candidates_xlsx(rows, path):
    """Write the week's candidates as a formatted, easy-to-review workbook:
    a dropdown on Include?, rows that highlight green once marked y, wrapped
    text so headlines are readable, a clickable Link column, and a frozen,
    filterable header row."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Candidates"

    n = len(REVIEW_COLS)
    for c, (key, label, width) in enumerate(REVIEW_COLS, start=1):
        cell = ws.cell(row=1, column=c, value=label)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        ws.column_dimensions[get_column_letter(c)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(n)}1"

    url_col = [k for k, _, _ in REVIEW_COLS].index("url") + 1
    for r, row in enumerate(rows, start=2):
        for c, (key, label, width) in enumerate(REVIEW_COLS, start=1):
            cell = ws.cell(row=r, column=c, value=row.get(key) or None)
            cell.font = BODY
            cell.alignment = WRAP
            cell.border = THIN
        if row.get("url"):
            link = ws.cell(row=r, column=url_col)
            link.hyperlink = row["url"]
            link.font = LINK_FONT

    last_row = max(len(rows) + 1, 2)

    dv = DataValidation(type="list", formula1='"y,n"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"A2:A{last_row}")

    green = PatternFill("solid", start_color=LIGHT_GREEN)
    ws.conditional_formatting.add(
        f"A2:{get_column_letter(n)}{last_row}",
        FormulaRule(formula=['LOWER($A2)="y"'], fill=green))

    wb.save(path)


def target_week(today=None):
    """Return (monday_date, start, end_exclusive) for the Mon–Sun week that
    contains YESTERDAY. On a Sunday run that is the week ending that Sunday."""
    today = today or dt.datetime.now(dt.timezone.utc).date()
    ref = today - dt.timedelta(days=1)          # yesterday
    monday = ref - dt.timedelta(days=ref.weekday())
    return monday, monday, monday + dt.timedelta(days=7)


def entry_date(e):
    for key in ("published_parsed", "updated_parsed"):
        t = getattr(e, key, None) or (e.get(key) if hasattr(e, "get") else None)
        if t:
            return dt.date(t.tm_year, t.tm_mon, t.tm_mday)
    return None


def relevant(title, summary):
    blob = (title + " " + summary).lower()
    return any(k in blob for k in KEYWORDS)


def collect(monday, start, end):
    seen_url, seen_title, rows = set(), set(), []
    week_label = monday.strftime("%b %-d, %Y")
    for outlet, url in FEEDS.items():
        try:
            feed = feedparser.parse(url)
        except Exception as ex:
            print(f"  ! skipped {outlet}: {ex}", file=sys.stderr)
            continue
        for e in feed.entries:
            d = entry_date(e)
            if not d or not (start <= d < end):
                continue
            title = re.sub(r"\s+", " ", (e.get("title") or "")).strip()
            summary = re.sub(r"<[^>]+>", " ", e.get("summary", "") or "")
            if not title or not relevant(title, summary):
                continue
            link = (e.get("link") or "").strip()
            norm = title.lower()
            if link in seen_url or norm in seen_title:
                continue
            seen_url.add(link); seen_title.add(norm)
            rows.append({
                "include": "", "week_of": week_label,
                "dates": d.strftime("%b %-d"), "category": "",
                "event": "", "impact": "", "outlet": outlet,
                "srcdesc": title, "url": link,
                "srcdate": d.strftime("%b %-d, %Y"),
            })
    rows.sort(key=lambda r: r["srcdate"])
    return rows[:MAX_CANDIDATES]


def ai_draft(rows):
    """Optional: pre-fill category/event/impact. Needs ANTHROPIC_API_KEY."""
    key = os.environ.get("ANTHROPIC_API_KEY")
    if not key:
        print("  (--draft set but no ANTHROPIC_API_KEY; leaving fields blank)")
        return rows
    try:
        import anthropic, json
    except ImportError:
        print("  (anthropic package not installed; leaving fields blank)")
        return rows
    client = anthropic.Anthropic(api_key=key)
    cats = ", ".join(CATEGORIES)
    for r in rows:
        prompt = (
            "You classify US news for a civil-liberties tracker of second-term "
            "Trump administration actions. Given one headline, respond ONLY with "
            "JSON: {\"category\":..., \"event\":..., \"impact\":...}. "
            f"category must be exactly one of: {cats}. "
            "event = <=25 words, plainly what the administration did. "
            "impact = <=40 words, why critics/courts/data call it harmful. "
            "If the headline is NOT about a specific administration action, set "
            "category to \"SKIP\" and leave event/impact empty.\n\n"
            f"Outlet: {r['outlet']}\nHeadline: {r['srcdesc']}"
        )
        try:
            msg = client.messages.create(
                model="claude-sonnet-4-6", max_tokens=400, temperature=0.2,
                messages=[{"role": "user", "content": prompt}])
            txt = "".join(b.text for b in msg.content if b.type == "text")
            txt = re.sub(r"^```json|```$", "", txt.strip()).strip()
            data = json.loads(txt)
            if data.get("category") and data["category"] != "SKIP":
                r["category"] = data.get("category", "")
                r["event"] = data.get("event", "")
                r["impact"] = data.get("impact", "")
        except Exception as ex:
            print(f"  ! draft failed for one item: {ex}", file=sys.stderr)
    return rows


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--draft", action="store_true",
                    help="Use Anthropic API to pre-fill category/event/impact")
    ap.add_argument("--outdir", default="candidates")
    args = ap.parse_args()

    monday, start, end = target_week()
    print(f"Gathering week of {monday} ({start} .. {end - dt.timedelta(days=1)})")
    rows = collect(monday, start, end)
    print(f"Collected {len(rows)} candidate items")
    if args.draft:
        rows = ai_draft(rows)

    os.makedirs(args.outdir, exist_ok=True)
    fname = f"{monday:%Y-%m-%d}.xlsx"
    path = os.path.join(args.outdir, fname)
    write_candidates_xlsx(rows, path)
    print(f"Wrote {path}")

    # expose to the workflow (for the review issue link/name)
    out = os.environ.get("GITHUB_OUTPUT")
    if out:
        with open(out, "a") as f:
            f.write(f"file={path}\n")
            f.write(f"count={len(rows)}\n")
            f.write(f"week={monday:%Y-%m-%d}\n")


if __name__ == "__main__":
    main()
