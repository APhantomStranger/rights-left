#!/usr/bin/env python3
"""
build_site.py — Stage 3 of the Rights Left pipeline (REBUILD SITE).

Reads the master workbook and regenerates rights-left.html in the repo root.
Called automatically by the ingest workflow after new rows are appended.

Usage:
    python scripts/build_site.py --xlsx data/Trump_Second_Term_Weekly_Tracker.xlsx
"""

import argparse, json, os, re, datetime as dt
from openpyxl import load_workbook

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_PATH = os.path.join(SCRIPT_DIR, "site_template.html")


def read_workbook(xlsx_path):
    """Extract timeline entries and source data from the workbook."""
    wb = load_workbook(xlsx_path, data_only=True)
    tl = wb["Weekly Timeline"]
    src = wb["Sources"]

    # Build a lookup of source data by ref number
    sources = {}
    for row in range(3, src.max_row + 1):
        ref = src.cell(row=row, column=1).value
        if ref is None:
            continue
        sources[ref] = {
            "outlet": src.cell(row=row, column=2).value or "",
            "srcdesc": src.cell(row=row, column=3).value or "",
            "url": src.cell(row=row, column=4).value or "",
            "srcdate": src.cell(row=row, column=5).value or "",
        }
        # openpyxl may return the hyperlink target instead of cell text for URLs
        hyp = src.cell(row=row, column=4).hyperlink
        if hyp and hyp.target:
            sources[ref]["url"] = hyp.target

    # Read timeline rows, group by week
    entries = []
    for row in range(3, tl.max_row + 1):
        week = tl.cell(row=row, column=1).value
        if week is None or str(week).strip() == "":
            continue
        ref = tl.cell(row=row, column=6).value
        src_data = sources.get(ref, {})
        entries.append({
            "week": str(week).strip(),
            "n": int(ref) if ref else row - 2,
            "date": str(tl.cell(row=row, column=2).value or "").strip(),
            "cat": str(tl.cell(row=row, column=3).value or "").strip(),
            "event": str(tl.cell(row=row, column=4).value or "").strip(),
            "impact": str(tl.cell(row=row, column=5).value or "").strip(),
            "outlet": src_data.get("outlet", ""),
            "srcdesc": src_data.get("srcdesc", ""),
            "url": src_data.get("url", ""),
            "srcdate": src_data.get("srcdate", ""),
            "image": "",
        })

    return entries


def _parse_week(week_label):
    """Parse a 'Week Of' label like 'Jul 13, 2026' into a real date.

    Falls back to datetime.min (sorts to the very end, i.e. oldest) for any
    label that doesn't match, so a formatting slip in the workbook can't
    crash the whole site build.
    """
    try:
        return dt.datetime.strptime(week_label, "%b %d, %Y")
    except ValueError:
        return dt.datetime.min


_MONTH_NUM = {m: i + 1 for i, m in enumerate(
    ["jan", "feb", "mar", "apr", "may", "jun",
     "jul", "aug", "sep", "oct", "nov", "dec"])}


def _parse_month_day(date_str):
    """Best-effort (month, day) from the free-text Date(s) column — 'Aug 20',
    'Jul 24-25', 'Feb 11+', even vague values like 'Mid-Feb'. Returns None
    for anything it can't confidently read.
    """
    s = date_str.strip().lower()
    m = re.match(r"(mid-)?([a-z]{3})[a-z]*\.?\s*(\d{1,2})?", s)
    if m and m.group(2) in _MONTH_NUM:
        month = _MONTH_NUM[m.group(2)]
        day = int(m.group(3)) if m.group(3) else (15 if m.group(1) else 1)
        return month, day
    return None


def _entry_date_key(entry, index):
    """Sort key for ordering entries within one week, newest first. See
    _parse_month_day — anything unparseable falls back to its original
    position instead of raising and breaking the whole site build.
    """
    md = _parse_month_day(entry["date"])
    if md:
        return (1, md[0], md[1], -index)
    return (0, 0, 0, -index)  # unparseable: sorts last, keeps original order


def _entry_date(entry, week_year):
    """Best-effort real date for an entry: its own month/day combined with
    the year from its 'Week Of' label. None if unparseable or invalid
    (e.g. a typo producing Feb 30).
    """
    md = _parse_month_day(entry["date"])
    if not md:
        return None
    try:
        return dt.date(week_year, md[0], md[1])
    except ValueError:
        return None


def group_by_week(entries):
    """Group entries into week objects, newest first.

    Entries are merged by week label first — NOT by adjacency in the sheet —
    so a week gets exactly one group even if rows for it were appended in
    more than one ingest run (e.g. a few rows approved early, more approved
    later). Groups are sorted by the actual parsed 'Week Of' date, newest
    first; entries *within* each week are likewise sorted newest first by
    their own 'Date(s)' value, so ordering is always correct regardless of
    what order rows happen to sit in on the Weekly Timeline sheet.
    """
    by_week = {}
    for e in entries:
        entry = {k: v for k, v in e.items() if k != "week"}
        by_week.setdefault(e["week"], []).append(entry)

    groups = []
    for w, es in by_week.items():
        keyed = sorted(enumerate(es), key=lambda pair: _entry_date_key(pair[1], pair[0]), reverse=True)
        groups.append({"week": w, "entries": [e for _, e in keyed]})

    groups.sort(key=lambda g: _parse_week(g["week"]), reverse=True)
    return groups


def build_site_json(groups, cats):
    """Build the SITE metadata object."""
    total = sum(len(g["entries"]) for g in groups)
    weeks = len(groups)
    # Date range: oldest week to today
    if groups:
        oldest = groups[-1]["week"]
        newest = groups[0]["week"]
    else:
        oldest = newest = "—"
    today = dt.date.today().strftime("%b %-d, %Y")
    return {
        "name": "Rights Left",
        "tag": "The Rights That Have Left of Us Project",
        "sub": ("Documented actions of the second Trump administration, "
                "logged week by week, with sources. Jan 20, 2025 – present."),
        "range": f"Jan 20, 2025 – {today}",
        "total": total,
        "weeks": weeks,
    }


def compute_stats(entries, today=None):
    """Category stats for the sidebar: the top 3 categories by entry count
    over the last 30 days ('trending'), and every category's all-time share
    ('totals', for the pie chart). Both skip entries still missing a
    category (not yet enriched) — an 'Uncategorized' slice would describe
    the pipeline's backlog, not the tracker's actual subject matter,
    which isn't what either chart is for. Recomputed fresh on every call,
    so both always reflect whatever's in the workbook as of this build.
    """
    today = today or dt.date.today()
    window_start = today - dt.timedelta(days=30)

    trend_counts, total_counts = {}, {}
    for e in entries:
        cat = e["cat"]
        if not cat:
            continue
        total_counts[cat] = total_counts.get(cat, 0) + 1

        parsed_week = _parse_week(e["week"])
        year = parsed_week.year if parsed_week != dt.datetime.min else today.year
        edate = _entry_date(e, year)
        if edate and window_start <= edate <= today:
            trend_counts[cat] = trend_counts.get(cat, 0) + 1

    trending = sorted(trend_counts.items(), key=lambda kv: kv[1], reverse=True)[:3]
    totals = sorted(total_counts.items(), key=lambda kv: kv[1], reverse=True)

    return {
        "trending": [{"category": c, "count": n} for c, n in trending],
        "totals": [{"category": c, "count": n} for c, n in totals],
    }


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True, help="Path to the master workbook")
    ap.add_argument("--out", default="index.html",
                    help="Output HTML path (default: index.html in repo root)")
    args = ap.parse_args()

    if not os.path.exists(TEMPLATE_PATH):
        raise FileNotFoundError(
            f"Template not found at {TEMPLATE_PATH}. "
            "It should be in the same directory as this script.")

    print(f"Reading workbook: {args.xlsx}")
    entries = read_workbook(args.xlsx)
    print(f"  {len(entries)} entries across {len(set(e['week'] for e in entries))} weeks")

    groups = group_by_week(entries)
    cats = sorted({e["cat"] for e in entries if e["cat"]})
    site = build_site_json(groups, cats)
    stats = compute_stats(entries)

    # Serialize to JSON, escaping </ for safe embedding in <script>
    data_json = json.dumps(groups, ensure_ascii=False).replace("</", "<\\/")
    cats_json = json.dumps(cats, ensure_ascii=False)
    site_json = json.dumps(site, ensure_ascii=False)
    stats_json = json.dumps(stats, ensure_ascii=False)

    template = open(TEMPLATE_PATH, encoding="utf-8").read()
    html = (template
            .replace("__DATA__", data_json)
            .replace("__CATS__", cats_json)
            .replace("__SITE__", site_json)
            .replace("__STATS__", stats_json))

    # Sanity check
    for token in ("__DATA__", "__CATS__", "__SITE__", "__STATS__"):
        if token in html:
            raise RuntimeError(f"Token {token} was not replaced — template issue")

    with open(args.out, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"Wrote {args.out} ({len(html):,} bytes)")
    print(f"  {site['total']} entries, {site['weeks']} weeks, {len(cats)} categories")


if __name__ == "__main__":
    main()
