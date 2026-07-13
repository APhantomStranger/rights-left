#!/usr/bin/env python3
"""
ingest.py — Stage 2 of the Rights Left pipeline (INGEST).

Reads an approved candidates CSV and appends the rows you marked include=y into
the existing workbook, reproducing the exact formatting already in the file:
 - Weekly Timeline: same Arial 10, wrap, hairline borders, per-week row banding,
   and continuing the File/Ref number sequence.
 - Sources: matching row with a clickable hyperlink.
 - Summary: rebuilt with whole-column COUNTIFs so new weeks/categories are
   always counted (no manual range edits ever needed).

The ONLY requirement to bring a row in is include=y. Every other field
(week_of, dates, category, event, impact, outlet, srcdesc, url, srcdate) is
optional — leave any of them blank and the cell is simply left empty in the
workbook. Fill them in later directly in the spreadsheet whenever you're
ready; nothing about the pipeline requires them to be complete at ingest time.

Nothing is fabricated here — it only formats and files what you approved.
Processed CSVs are moved to candidates/processed/ so they can't be ingested twice.
"""

import csv, os, sys, glob, shutil, argparse
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# --- styling constants copied verbatim from the original build ----------------
NAVY, LIGHT = "1F3864", "D9E2F3"
BODY = Font(name="Arial", size=10)
LINK = Font(name="Arial", size=10, color="0563C1", underline="single")
HDR = Font(name="Arial", size=10, bold=True, color="FFFFFF")
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Border(bottom=Side(style="thin", color="BFBFBF"))
TRUTHY = {"y", "yes", "true", "1", "x"}


def is_light(cell):
    try:
        rgb = cell.fill.start_color.rgb
        return isinstance(rgb, str) and rgb.upper().endswith(LIGHT)
    except Exception:
        return False


def last_data_row(ws):
    r = 2
    for row in range(3, ws.max_row + 1):
        if ws.cell(row=row, column=1).value not in (None, ""):
            r = row
    return r


def read_approved(path):
    """Return every row marked include=y. No other field is required —
    category/event/impact/outlet/etc. can be filled in later directly in the
    workbook. A row with everything blank except include=y still gets a File
    No. and a placeholder position in the timeline; you edit the cells after."""
    rows, incomplete = [], 0
    with open(path, newline="", encoding="utf-8") as f:
        for r in csv.DictReader(f):
            if (r.get("include") or "").strip().lower() not in TRUTHY:
                continue
            row = {k: (r.get(k) or "").strip() for k in r}
            blanks = [k for k in ("week_of", "dates", "category", "event",
                                  "impact", "outlet", "srcdesc")
                      if not row.get(k)]
            if blanks:
                incomplete += 1
                print(f"  · included with blank fields ({', '.join(blanks)}): "
                      f"{(row.get('srcdesc') or row.get('event') or '(no title)')[:60]}")
            rows.append(row)
    if incomplete:
        print(f"  {incomplete} row(s) ingested with some fields left blank — "
              f"fill those in directly in the workbook whenever you're ready.")
    return rows


def rebuild_summary(wb, ws_time_name, timeline_last):
    ws = wb["Summary"]
    # collect categories present in the timeline
    tl = wb[ws_time_name]
    cats = sorted({tl.cell(row=r, column=3).value
                   for r in range(3, timeline_last + 1)
                   if tl.cell(row=r, column=3).value})
    # clear any existing category rows (row 3 downward, cols A & B)
    for r in range(3, ws.max_row + 1):
        ws.cell(row=r, column=1).value = None
        ws.cell(row=r, column=2).value = None
    # write fresh, using whole-column COUNTIF so ranges never go stale
    for i, cat in enumerate(cats, start=3):
        ws.cell(row=i, column=1, value=cat).font = BODY
        c = ws.cell(row=i, column=2,
                    value=f"=COUNTIF('{ws_time_name}'!C:C,A{i})")
        c.font = BODY
    tot = 3 + len(cats)
    a = ws.cell(row=tot, column=1, value="Total entries")
    b = ws.cell(row=tot, column=2, value=f"=SUM(B3:B{tot - 1})")
    a.font = Font(name="Arial", size=10, bold=True)
    b.font = Font(name="Arial", size=10, bold=True)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True)
    ap.add_argument("--csv", help="approved CSV (default: newest in candidates/)")
    ap.add_argument("--candir", default="candidates")
    ap.add_argument("--recalc", action="store_true",
                    help="round-trip through LibreOffice to bake formula values")
    args = ap.parse_args()

    csv_path = args.csv
    if not csv_path:
        pool = sorted(glob.glob(os.path.join(args.candir, "*.csv")))
        if not pool:
            sys.exit("No candidates CSV found to ingest.")
        csv_path = pool[-1]
    print(f"Ingesting: {csv_path}")

    approved = read_approved(csv_path)
    if not approved:
        print("No rows marked include=y — nothing to ingest.")
        # still archive the file so it isn't reprocessed
        _archive(csv_path, args.candir)
        return

    wb = load_workbook(args.xlsx)
    for need in ("Weekly Timeline", "Sources", "Summary"):
        if need not in wb.sheetnames:
            sys.exit(f"Workbook is missing the '{need}' sheet — wrong file?")
    tl, src = wb["Weekly Timeline"], wb["Sources"]

    # existing state
    tl_last = last_data_row(tl)
    src_last = last_data_row(src)
    max_ref = 0
    for r in range(3, tl_last + 1):
        v = tl.cell(row=r, column=6).value
        if isinstance(v, int):
            max_ref = max(max_ref, v)
    existing_urls = {src.cell(row=r, column=4).value
                     for r in range(3, src_last + 1)
                     if src.cell(row=r, column=4).value}

    band = is_light(tl.cell(row=tl_last, column=1))
    prev_week = tl.cell(row=tl_last, column=1).value

    added = 0
    tr, sr = tl_last, src_last
    for row in approved:
        if row.get("url") and row["url"] in existing_urls:
            print(f"  · duplicate URL, skipping: {row['url']}")
            continue
        max_ref += 1
        if row["week_of"] != prev_week:
            band = not band
            prev_week = row["week_of"]

        tr += 1
        vals = [row["week_of"] or None, row["dates"] or None,
                row["category"] or None, row["event"] or None,
                row["impact"] or None, max_ref]
        for c, v in enumerate(vals, start=1):
            cell = tl.cell(row=tr, column=c, value=v)
            cell.font = BODY
            cell.alignment = WRAP
            cell.border = THIN
            if band:
                cell.fill = PatternFill("solid", start_color=LIGHT)

        sr += 1
        svals = [max_ref, row["outlet"] or None, row["srcdesc"] or None,
                 row["url"] or None, row["srcdate"] or None]
        for c, v in enumerate(svals, start=1):
            cell = src.cell(row=sr, column=c, value=v)
            cell.font = BODY
            cell.alignment = WRAP
            cell.border = THIN
        if row.get("url"):
            link = src.cell(row=sr, column=4)
            link.hyperlink = row["url"]
            link.font = LINK
        if row.get("url"):
            existing_urls.add(row["url"])
        added += 1

    # keep the autofilter covering all rows
    tl.auto_filter.ref = f"A2:F{tr}"
    rebuild_summary(wb, "Weekly Timeline", tr)

    wb.save(args.xlsx)
    print(f"Appended {added} row(s). Timeline now ends at row {tr}, "
          f"last File No. {max_ref}.")

    if args.recalc:
        _recalc(args.xlsx)
    _archive(csv_path, args.candir)


def _archive(csv_path, candir):
    done = os.path.join(candir, "processed")
    os.makedirs(done, exist_ok=True)
    dest = os.path.join(done, os.path.basename(csv_path))
    try:
        shutil.move(csv_path, dest)
        print(f"Moved {csv_path} -> {dest}")
    except Exception as ex:
        print(f"  (could not archive CSV: {ex})")


def _recalc(xlsx):
    """Optional: force LibreOffice to recompute + cache formula values so the
    committed file shows correct numbers even before it's opened in Excel."""
    import subprocess, tempfile
    try:
        with tempfile.TemporaryDirectory() as td:
            subprocess.run(
                ["soffice", "--headless", "--calc", "--convert-to", "xlsx",
                 "--outdir", td, xlsx],
                check=True, timeout=120,
                stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            out = os.path.join(td, os.path.basename(xlsx))
            if os.path.exists(out):
                shutil.copy(out, xlsx)
                print("Recalculated formula values via LibreOffice.")
    except Exception as ex:
        print(f"  (recalc skipped — Excel will recompute on open: {ex})")


if __name__ == "__main__":
    main()
